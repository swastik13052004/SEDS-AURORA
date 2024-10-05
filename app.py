from flask import Flask, render_template, request, redirect, url_for, flash
import numpy as np
import pickle
import openpyxl
import os
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Set a secret key for flash messages
model = pickle.load(open('Kidney.pkl', 'rb'))

# Define the Excel file paths
user_data_file = 'user_data.xlsx'
prediction_data_file = 'prediction_data.xlsx'

# Initialize the Excel files with headers if they don't exist
def initialize_excel_file(file_path, headers):
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(headers)
        workbook.save(file_path)

initialize_excel_file(user_data_file, ['Email', 'Password'])
initialize_excel_file(prediction_data_file, ['Specific Gravity', 'Hypertension', 'Hemoglobin', 'Diabetes Mellitus', 'Albumin', 'Appetite', 'Red Blood Cells', 'Pus Cell'])

@app.route('/', methods=['GET', 'POST'])
def home():
    if request.method == 'POST':
        email = request.form['loginEmail']
        password = request.form['loginPassword']
        
        # Check user credentials
        workbook = openpyxl.load_workbook(user_data_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == email and check_password_hash(row[1], password):
                flash('Login successful!', 'success')
                return redirect(url_for('index'))
        
        flash('Invalid email or password', 'error')
    
    return render_template('login_signup.html')  # Updated to use login_signup.html

@app.route('/signup', methods=['POST'])
def signup():
    name = request.form['signupName']
    email = request.form['signupEmail']
    password = request.form['signupPassword']
    confirm_password = request.form['signupConfirmPassword']

    if password != confirm_password:
        flash('Passwords do not match', 'error')
        return redirect(url_for('home'))

    # Check if email already exists
    workbook = openpyxl.load_workbook(user_data_file)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == email:
            flash('Email already in use', 'error')
            return redirect(url_for('home'))

    # Add new user
    hashed_password = generate_password_hash(password)
    sheet.append([email, hashed_password])
    workbook.save(user_data_file)

    flash('Signup successful! You can now log in.', 'success')
    return redirect(url_for('home'))

@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/contact')
def contact():
    return render_template('contact.html')

@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username == 'admin' and password == 'admin':
            # Load the Excel file and retrieve user data
            workbook = openpyxl.load_workbook(prediction_data_file)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            return render_template('admin_data.html', data=data)
        else:
            flash('Invalid admin credentials', 'error')
    
    return render_template('admin.html')

@app.route("/predict", methods=['POST'])
def predict():
    if request.method == 'POST':
        sg = float(request.form['sg'])
        htn = float(request.form['htn'])
        hemo = float(request.form['hemo'])
        dm = float(request.form['dm'])
        al = float(request.form['al'])
        appet = float(request.form['appet'])
        rc = float(request.form['rc'])
        pc = float(request.form['pc'])

        # Save the user input to the Excel file
        workbook = openpyxl.load_workbook(prediction_data_file)
        sheet = workbook.active
        sheet.append([sg, htn, hemo, dm, al, appet, rc, pc])
        workbook.save(prediction_data_file)

        # Predict using the model
        values = np.array([[sg, htn, hemo, dm, al, appet, rc, pc]])
        prediction = model.predict(values)

        return render_template('result.html', prediction=prediction)

if __name__ == "__main__":
    app.run(debug=True)